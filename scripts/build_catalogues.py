#!/usr/bin/env python3
"""Turn 119 MB of published sources into the catalogues the method reads.

Two kinds of output, both committed so the test suite never needs the network:

  catalogues/items/{framework}.json   id, title and text per item
  catalogues/pairs/{a}__{b}.json      the human-made mappings, normalised

Every structure has its size asserted. On the Regulatory Scope Ledger that
discipline caught a claim of 20 DORA entity categories when there are 21; here
it is the only thing standing between a silent parsing failure and published
numbers that mean nothing.

**Identifiers are normalised, and that is not cosmetic.** CTID writes a
control as `AC-02`, OSCAL writes it `ac-2`, D3FEND writes it `CM-5(3)`. The
raw intersection of CTID's controls with OSCAL's is *zero*. After
normalisation it is 109 of 109. A pipeline without this step would report a
method that recovers nothing, and the fault would look like the method's.
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Iterable
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CACHE = ROOT / ".cache" / "sources"
ITEMS = ROOT / "catalogues" / "items"
PAIRS = ROOT / "catalogues" / "pairs"


# --------------------------------------------------------------------------
# Identifiers
# --------------------------------------------------------------------------


def control_id(raw: str) -> str:
    """Canonical 800-53 control id.

    `AC-02` → `ac-2`, `CM-5(3)` → `cm-5.3`, `ac-2.1` → `ac-2.1`.
    """
    text = raw.strip().lower().replace(" ", "")
    text = re.sub(r"\((\d+)\)", r".\1", text)
    # Zero-padding is CTID's convention and nobody else's.
    text = re.sub(r"^([a-z]{2})-0*(\d+)", r"\1-\2", text)
    return re.sub(r"\.0*(\d+)", r".\1", text)


def technique_id(raw: str) -> str:
    return raw.strip().upper()


def normalise_text(raw: str) -> str:
    """NFKC, then collapse whitespace.

    The non-breaking space is why this exists. The DORA source carries 1,650
    of them, and a regex written with a plain space matches none of the
    article headings. That failure is silent: you get zero articles and a
    catalogue that looks merely empty.
    """
    text = unicodedata.normalize("NFKC", raw).replace(" ", " ")
    return re.sub(r"\s+", " ", text).strip()


def strip_tags(raw: str) -> str:
    return normalise_text(re.sub(r"<[^>]+>", " ", raw))


# --------------------------------------------------------------------------
# Items
# --------------------------------------------------------------------------


def build_attack() -> list[dict]:
    data = json.loads((CACHE / "attack-enterprise.json").read_text(encoding="utf-8"))
    items = []
    for obj in data["objects"]:
        if obj["type"] != "attack-pattern" or obj.get("revoked") or obj.get("x_mitre_deprecated"):
            continue
        ref = next(
            (
                r
                for r in obj.get("external_references", [])
                if r.get("source_name") == "mitre-attack"
            ),
            None,
        )
        if not ref:
            continue
        tactics = [p["phase_name"] for p in obj.get("kill_chain_phases", [])]
        items.append(
            {
                "id": technique_id(ref["external_id"]),
                "title": obj["name"],
                "text": normalise_text(obj.get("description", "")),
                "group": ",".join(sorted(tactics)),
                "is_subtechnique": bool(obj.get("x_mitre_is_subtechnique")),
            }
        )
    assert len(items) == 656, f"ATT&CK 16.1 should hold 656 live techniques, found {len(items)}"
    return sorted(items, key=lambda i: i["id"])


def _oscal_prose(part: dict) -> Iterable[str]:
    if part.get("prose"):
        yield part["prose"]
    for child in part.get("parts", []):
        yield from _oscal_prose(child)


#: OSCAL leaves parameter slots in the prose. Resolving them needs a profile
#: and a baseline; this project only needs the words, so the slot is replaced
#: by a neutral placeholder rather than deleted, which would join two
#: sentences into nonsense.
_PARAM = re.compile(r"\{\{\s*insert:\s*param,\s*[^}]+\}\}")


def build_800_53() -> list[dict]:
    catalog = json.loads((CACHE / "nist-800-53r5-catalog.json").read_text(encoding="utf-8"))[
        "catalog"
    ]
    items: list[dict] = []

    def visit(group: dict, family: str) -> None:
        for control in group.get("controls", []):
            _add(control, family)
        for child in group.get("groups", []):
            visit(child, child.get("title", family))

    withdrawn: set[str] = set()

    def _add(control: dict, family: str) -> None:
        # Rev 5 keeps withdrawn controls in the catalogue as tombstones: a
        # title, no statement, status 'withdrawn'. They carry no text to
        # embed and must never be offered as candidates, so they are recorded
        # and excluded rather than counted as a parsing failure. Discovered
        # because the empty-text assertion fired on 180 of them.
        status = next(
            (p["value"] for p in control.get("props", []) if p.get("name") == "status"), ""
        )
        canonical = control_id(control["id"])
        if status == "withdrawn":
            withdrawn.add(canonical)
        else:
            statement = next(
                (p for p in control.get("parts", []) if p.get("name") == "statement"), {}
            )
            prose = " ".join(_oscal_prose(statement)) if statement else ""
            items.append(
                {
                    "id": canonical,
                    "title": control["title"],
                    "text": normalise_text(_PARAM.sub("[parameter]", prose)),
                    "group": family,
                    "is_enhancement": "." in control["id"],
                }
            )
        for child in control.get("controls", []):
            _add(child, family)

    for group in catalog["groups"]:
        visit(group, group.get("title", group["id"]))

    total = len(items) + len(withdrawn)
    assert total == 1196, f"800-53 Rev 5 should hold 1196 entries, found {total}"
    assert len(withdrawn) == 182, f"expected 182 withdrawn controls, found {len(withdrawn)}"
    empty = [i["id"] for i in items if not i["text"]]
    assert not empty, f"{len(empty)} live controls parsed with no statement text: {empty[:8]}"
    (ITEMS.parent / "nist-800-53-withdrawn.json").parent.mkdir(parents=True, exist_ok=True)
    (ITEMS.parent / "nist-800-53-withdrawn.json").write_text(
        json.dumps(sorted(withdrawn), indent=1) + "\n", encoding="utf-8"
    )
    return sorted(items, key=lambda i: i["id"])


def _xlsx_rows(path: Path, sheet_name: str) -> list[list[str]]:
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    sheet = book[sheet_name]
    rows = [
        ["" if cell is None else str(cell).strip() for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]
    book.close()
    return rows


_SUBCAT = re.compile(r"^([A-Z]{2}\.[A-Z]{2}-\d+)")


def build_csf_and_pairs() -> tuple[list[dict], list[dict]]:
    """CSF 1.1 subcategories, and the NIST crosswalk to 800-53.

    This is version 1.1, not 2.0. The file name does not say so; the title
    inside does, and the absence of any GV.* identifier proves it — the GOVERN
    function exists only in 2.0. NIST moved its 2.0 references to the OLIR
    catalogue, which has no download endpoint.
    """
    rows = _xlsx_rows(CACHE / "csf-pf-to-800-53r5.xlsx", "CSF to SP 800-53r5")

    header = rows[1]
    assert header[:4] == [
        "Function",
        "Category",
        "Subcategory",
        "NIST SP 800-53, Revision 5 Control",
    ], f"the sheet layout changed: {header[:4]}"

    items: dict[str, dict] = {}
    unique: set[tuple[str, str]] = set()
    function = ""

    for row in rows[2:]:
        if row and row[0]:
            function = row[0].split("(")[0].strip()
        cell = row[2] if len(row) > 2 else ""
        match = re.match(r"^([A-Z]{2}\.[A-Z]{2}-\d+)\s*:\s*(.*)$", cell, re.S)
        if not match:
            continue
        sub_id, text = match.group(1), normalise_text(match.group(2))
        if sub_id.startswith("GV."):
            raise AssertionError("a GOVERN identifier appeared: this file is not CSF 1.1")

        items[sub_id] = {"id": sub_id, "title": sub_id, "text": text, "group": function}
        for control in re.findall(
            r"\b([A-Z]{2}-\d{1,2}(?:\(\d+\))?)\b", row[3] if len(row) > 3 else ""
        ):
            unique.add((sub_id, control_id(control)))

    assert len(items) == 108, f"CSF 1.1 holds 108 subcategories, found {len(items)}"
    # 495 verified against the sheet: 108 subcategories, 4.6 controls each on
    # average, and ID.GV-1 and ID.GV-3 mapped to nothing at all.
    assert len(unique) == 495, f"expected 495 CSF↔800-53 pairs, parsed {len(unique)}"
    assert all(i["text"] for i in items.values()), "a subcategory parsed with no text"
    return sorted(items.values(), key=lambda i: i["id"]), [
        {"left": sub, "right": ctl, "source": "nist-csf1.1"} for sub, ctl in sorted(unique)
    ]


def build_d3fend() -> tuple[list[dict], list[dict]]:
    """D3FEND defensive techniques, and their mapping to 800-53 controls."""
    graph = json.loads((CACHE / "d3fend.json").read_text(encoding="utf-8"))["@graph"]
    by_id = {node["@id"]: node for node in graph if "@id" in node}

    def types_of(node: dict) -> list[str]:
        value = node.get("@type")
        return [value] if isinstance(value, str) else (value or [])

    def label(node: dict) -> str:
        value = node.get("rdfs:label", "")
        return value if isinstance(value, str) else (value or {}).get("@value", "")

    # Defensive techniques are not typed as such: they are owl:Class nodes
    # carrying a d3f:d3fend-id. The prefix separates them from the analytic
    # vocabulary, which uses D3A- and is not a countermeasure.
    items = []
    for node in graph:
        identifier = str(node.get("d3f:d3fend-id", ""))
        if not identifier.startswith("D3-"):
            continue
        text = node.get("d3f:definition") or node.get("rdfs:comment") or ""
        if isinstance(text, dict):
            text = text.get("@value", "")
        name = label(node)
        if not name:
            continue
        items.append(
            {
                "id": node["@id"].removeprefix("d3f:"),
                "title": name,
                "text": normalise_text(str(text)),
                "group": identifier,
            }
        )

    pairs: list[dict] = []
    for node in graph:
        if "d3f:NISTControl" not in types_of(node):
            continue
        control = label(node)
        if not control:
            continue
        narrower = node.get("d3f:narrower", [])
        if isinstance(narrower, dict):
            narrower = [narrower]
        for target in narrower:
            ref = by_id.get(target.get("@id"), {})
            if str(ref.get("d3f:d3fend-id", "")).startswith("D3-"):
                pairs.append(
                    {
                        "left": target["@id"].removeprefix("d3f:"),
                        "right": control_id(control),
                        "source": "d3fend",
                    }
                )

    assert len(items) == 272, f"D3FEND holds 272 defensive techniques, found {len(items)}"
    assert all(i["text"] for i in items), "a D3FEND technique parsed with no definition"
    unique = {(p["left"], p["right"]) for p in pairs}
    # Only 110, from 117 NISTControl nodes. This is by far the thinnest of the
    # four ground-truth pairs and any recall measured on it carries very wide
    # error bars. Recorded here so the number is not read as if it were the
    # 5,314-pair CTID set.
    assert len(unique) == 110, f"expected 110 D3FEND↔800-53 pairs, parsed {len(unique)}"
    return sorted(items, key=lambda i: i["id"]), [
        {"left": tech, "right": ctl, "source": "d3fend"} for tech, ctl in sorted(unique)
    ]


def build_attack_d3fend_pairs() -> list[dict]:
    rows = json.loads((CACHE / "d3fend-full-mappings.json").read_text(encoding="utf-8"))
    unique = set()
    for row in rows["results"]["bindings"]:
        offensive = row.get("off_tech_id", {}).get("value", "")
        defensive = row.get("def_tech", {}).get("value", "")
        if not offensive or not defensive:
            continue
        # off_tech_id also carries detection identifiers (DE-0002, DE-0003.08),
        # which arrived with ATT&CK 16 and are not techniques. Left in, they
        # became 104 phantom counterparts the method could never match.
        if not re.fullmatch(r"T\d{4}(\.\d{3})?", offensive):
            continue
        # T0xxx is ICS ATT&CK. D3FEND maps into it; this project is scoped to
        # Enterprise for v1, so those pairs are out of scope rather than
        # broken, and are filtered here instead of surfacing as 725 dangling
        # rows that look like a parsing failure.
        if offensive.startswith("T0"):
            continue
        unique.add((technique_id(offensive), defensive.rsplit("/", 1)[-1].rsplit("#", 1)[-1]))
    assert len(unique) > 1000, f"only {len(unique)} ATT&CK↔D3FEND pairs parsed"
    return [{"left": a, "right": d, "source": "d3fend"} for a, d in sorted(unique)]


def build_ctid_pairs() -> list[dict]:
    rows = json.loads((CACHE / "ctid-800-53r5-to-attack.json").read_text(encoding="utf-8"))
    unique = {
        (control_id(row["capability_id"]), technique_id(row["attack_object_id"]))
        for row in rows["mapping_objects"]
        if row.get("capability_id") and row.get("mapping_type") == "mitigates"
    }
    assert len(unique) > 5000, f"only {len(unique)} CTID pairs parsed"
    return [{"left": c, "right": t, "source": "ctid"} for c, t in sorted(unique)]


_ARTICLE = re.compile(r"^Article\s+(\d{1,2})$")


def build_dora() -> list[dict]:
    """DORA articles, as obligations to be matched against controls.

    Recitals are excluded: they explain intent and are not obligations, and
    including them would inflate every similarity score with restatements of
    the articles.
    """
    raw = (CACHE / "dora.xhtml").read_text(encoding="utf-8")
    blocks = [strip_tags(b) for b in re.split(r"</?(?:p|div|h[1-6])[^>]*>", raw)]
    blocks = [b for b in blocks if b]

    articles: list[dict] = []
    index = 0
    while index < len(blocks):
        match = _ARTICLE.match(blocks[index])
        if not match:
            index += 1
            continue
        number = int(match.group(1))
        title = blocks[index + 1] if index + 1 < len(blocks) else ""
        body: list[str] = []
        index += 2
        while index < len(blocks) and not _ARTICLE.match(blocks[index]):
            body.append(blocks[index])
            index += 1
        text = normalise_text(" ".join(body))
        if len(text) < 80:
            continue
        articles.append(
            {
                "id": f"Article {number}",
                "title": title,
                "text": text,
                "group": "",
            }
        )

    seen: dict[str, dict] = {}
    for article in articles:
        if article["id"] not in seen or len(article["text"]) > len(seen[article["id"]]["text"]):
            seen[article["id"]] = article

    assert 55 <= len(seen) <= 64, f"DORA has 64 articles, parsed {len(seen)}"
    return sorted(seen.values(), key=lambda a: int(a["id"].split()[1]))


# --------------------------------------------------------------------------


def write(folder: Path, name: str, payload: list[dict]) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    (folder / f"{name}.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )
    print(f"  {name:<28} {len(payload):>6,}")


def main() -> int:
    print("items")
    attack = build_attack()
    write(ITEMS, "attack", attack)
    controls = build_800_53()
    write(ITEMS, "nist-800-53", controls)
    csf, csf_pairs = build_csf_and_pairs()
    write(ITEMS, "nist-csf-1.1", csf)
    d3fend, d3f_pairs = build_d3fend()
    write(ITEMS, "d3fend", d3fend)
    dora = build_dora()
    write(ITEMS, "dora", dora)

    print("\npairs (ground truth)")
    write(PAIRS, "nist-800-53__attack", build_ctid_pairs())
    write(PAIRS, "attack__d3fend", build_attack_d3fend_pairs())
    write(PAIRS, "d3fend__nist-800-53", d3f_pairs)
    write(PAIRS, "nist-csf-1.1__nist-800-53", csf_pairs)

    # Every identifier in a pair must exist in its catalogue, or recall is
    # measured against counterparts the method was never shown.
    print("\nreferential integrity")
    known = {
        "attack": {i["id"] for i in attack},
        "nist-800-53": {i["id"] for i in controls},
        "nist-csf-1.1": {i["id"] for i in csf},
        "d3fend": {i["id"] for i in d3fend},
    }
    # A pair whose endpoint is not in its catalogue is dropped, because recall
    # cannot be measured against a counterpart the method was never shown. It
    # is recorded rather than discarded quietly: what a published crosswalk
    # points at and the framework no longer contains is itself a finding.
    withdrawn = set(json.loads((ROOT / "catalogues" / "nist-800-53-withdrawn.json").read_text()))
    dangling: dict[str, dict] = {}
    problems = 0

    for path in sorted(PAIRS.glob("*.json")):
        left_name, right_name = path.stem.split("__")
        rows = json.loads(path.read_text(encoding="utf-8"))
        bad_l = {r["left"] for r in rows} - known[left_name]
        bad_r = {r["right"] for r in rows} - known[right_name]

        kept = [r for r in rows if r["left"] not in bad_l and r["right"] not in bad_r]
        if len(kept) != len(rows):
            path.write_text(json.dumps(kept, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
            dangling[path.stem] = {
                "dropped_rows": len(rows) - len(kept),
                "left_not_in_catalogue": sorted(bad_l),
                "right_not_in_catalogue": sorted(bad_r),
                "right_withdrawn_in_rev5": sorted(bad_r & withdrawn),
            }

        note = ""
        if bad_r & withdrawn:
            note = f"  ({len(bad_r & withdrawn)} withdrawn in Rev 5)"
        flag = "  " if not (bad_l or bad_r) else "· "
        print(
            f"{flag} {path.stem:<30} kept {len(kept):>5,}  dropped {len(rows) - len(kept):>3}{note}"
        )
        # Losing more than a twentieth of a mapping means the parse is wrong,
        # not that the publishers disagree at the edges. Controls the
        # publisher withdrew are that ordinary disagreement, so they do not
        # count towards the limit.
        unexplained = (len(rows) - len(kept)) - len(
            [r for r in rows if r["right"] in (bad_r & withdrawn)]
        )
        if rows and unexplained / len(rows) > 0.05:
            print(f"     TOO MANY: left e.g. {sorted(bad_l)[:4]} right e.g. {sorted(bad_r)[:4]}")
            problems += 1

    (ROOT / "catalogues" / "dangling.json").write_text(
        json.dumps(dangling, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )

    print(f"\nDORA articles parsed: {len(dora)}")
    for name, info in dangling.items():
        print(f"  dangling in {name}: {info['dropped_rows']} rows")
    return 1 if problems else 0


if __name__ == "__main__":
    raise SystemExit(main())
