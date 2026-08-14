#!/usr/bin/env python3
"""Step 3: candidate crosswalk from DORA obligations to control frameworks.

Every row is a **candidate for a person to accept or reject**. Nothing here is
a mapping, and the word is avoided deliberately: no official DORA crosswalk
exists, so there is nothing to be right about.

Three things are attached to the output so a reader cannot take it for more
than it is:

  * every row is labelled `proposed`, because none can be `official` or
    `composed` — there is no published DORA mapping to compose from;
  * the measured performance of this method on comparable pairs travels with
    the file, from reports/retrieval.json;
  * so does the human-agreement figure, because two experts on a comparable
    task agreed on 19% of their pairs, which is the context that stops any
    single number here being read as accuracy.

**DORA has never been measured and cannot be.** The recall figures in
reports/retrieval.json come from 800-53, CSF, ATT&CK and D3FEND. There is no
human DORA crosswalk to score against, which is why this project exists and
also why these candidates carry no accuracy claim of their own.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import numpy as np

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from crosswalk_ledger.obligations import split_all  # noqa: E402

ITEMS = ROOT / "catalogues" / "items"
REPORTS = ROOT / "reports"
HUMAN = ROOT / "human-mappings"
MODEL = "sentence-transformers/all-MiniLM-L6-v2"

#: How many candidates to keep per obligation. Three because that is the
#: density two human mappers worked at, and the only cut-off for which this
#: project has a comparable human figure.
TOP_N = 3

#: ATT&CK is deliberately not a direct target. DORA states duties on
#: defenders; ATT&CK describes what attackers do. They are not the same kind
#: of statement, and similarity over text cannot bridge that — it matches
#: vocabulary and returns confident nonsense. The first run proposed
#: Article 45, "information-sharing arrangements on cyber threat
#: information", against T1597.002 "Purchase Technical Data", which is an
#: adversary buying data, and Article 35, "Powers of the Lead Overseer",
#: against T1199 "Trusted Relationship".
#:
#: The route to ATT&CK is through a control catalogue, where each link
#: compares like with like: DORA to 800-53 by similarity, then 800-53 to
#: ATT&CK by MITRE's published crosswalk.
TARGETS = ("nist-800-53", "nist-csf-2.0")


def load(name: str) -> list[dict]:
    return json.loads((ITEMS / f"{name}.json").read_text(encoding="utf-8"))


def build_csf_2() -> list[dict]:
    """CSF 2.0 subcategories and their text.

    Taken from the OLIR submission spreadsheet, which carries the focal
    document's own descriptions. The CSF text is NIST's and free to reuse; it
    is the ISO column of that same file that cannot be redistributed, and it
    is empty here in any case.
    """
    from openpyxl import load_workbook

    book = load_workbook(HUMAN / "razilio-csf2.0-to-iso27001.xlsx", read_only=True, data_only=True)
    sheet = book[book.sheetnames[0]]
    items: dict[str, dict] = {}
    for row in sheet.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if len(cells) < 2 or not re.fullmatch(r"[A-Z]{2}\.[A-Z]{2}-\d{2}", cells[0]):
            continue
        if cells[1]:
            items[cells[0]] = {
                "id": cells[0],
                "title": cells[0],
                "text": cells[1],
                "group": cells[0].split(".")[0],
            }
    book.close()
    assert len(items) == 106, f"CSF 2.0 holds 106 subcategories, parsed {len(items)}"
    ordered = sorted(items.values(), key=lambda i: i["id"])
    (ITEMS / "nist-csf-2.0.json").write_text(
        json.dumps(ordered, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )
    return ordered


def encode(texts: list[str]) -> np.ndarray:
    from sentence_transformers import SentenceTransformer

    model = SentenceTransformer(MODEL)
    return np.asarray(model.encode(texts, normalize_embeddings=True, show_progress_bar=False))


def main() -> int:
    obligations = split_all(load("dora"))
    print(f"DORA obligations: {len(obligations)}")

    build_csf_2()
    catalogues = {name: load(name) for name in TARGETS}
    for name, items in catalogues.items():
        print(f"  {name}: {len(items)} candidates")

    left = encode([f"{o['title']}. {o['text']}" for o in obligations])

    rows: list[dict] = []
    for name, items in catalogues.items():
        right = encode([f"{i['title']}. {i['text']}" for i in items])
        scores = left @ right.T
        for row_index, obligation in enumerate(obligations):
            ranked = np.argsort(-scores[row_index])[:TOP_N]
            for rank, column in enumerate(ranked, start=1):
                rows.append(
                    {
                        "dora": obligation["id"],
                        "dora_article": obligation["group"],
                        "dora_title": obligation["title"],
                        "target_framework": name,
                        "target": items[column]["id"],
                        "target_title": items[column]["title"],
                        "rank": rank,
                        "similarity": round(float(scores[row_index][column]), 4),
                        "provenance": "proposed",
                        "accepted_by": "",
                        "accepted_at": "",
                    }
                )

    # Compose the second leg from MITRE's published mapping rather than
    # measuring DORA against ATT&CK directly.
    ctid = json.loads(
        (ROOT / "catalogues" / "pairs" / "nist-800-53__attack.json").read_text(encoding="utf-8")
    )
    control_to_attack: dict[str, list[str]] = {}
    for pair in ctid:
        control_to_attack.setdefault(pair["left"], []).append(pair["right"])

    composed = []
    for row in rows:
        if row["target_framework"] != "nist-800-53":
            continue
        for technique in control_to_attack.get(row["target"], []):
            composed.append(
                {
                    "dora": row["dora"],
                    "dora_article": row["dora_article"],
                    "via_control": row["target"],
                    "attack": technique,
                    "first_leg": "proposed",
                    "second_leg": "official (MITRE CTID)",
                    "similarity_of_first_leg": row["similarity"],
                }
            )
    print(
        f"composed DORA → 800-53 → ATT&CK: {len(composed)} rows "
        f"via {len({c['via_control'] for c in composed})} controls"
    )

    retrieval = json.loads((REPORTS / "retrieval.json").read_text(encoding="utf-8"))
    agreement = json.loads((REPORTS / "agreement.json").read_text(encoding="utf-8"))

    payload = {
        "what_this_is": (
            "Candidate correspondences between DORA obligations and control frameworks, "
            "produced by embedding similarity. Every row is a proposal for a person to "
            "accept or reject. None of them is a mapping."
        ),
        "model": MODEL,
        "top_n": TOP_N,
        "dora_obligations": len(obligations),
        "dora_scope": "Articles 5-45. Definitions, scope, supervision and amendments excluded.",
        "no_dora_measurement_exists": (
            "No human DORA crosswalk has been published, so these candidates have not been "
            "and cannot be scored. The recall figures below come from other framework pairs "
            "and describe the method, not this output."
        ),
        "method_performance_on_other_pairs": {
            r["direction"]: r["recall_at"]
            for r in retrieval["results"]
            if r["method"] == "embedding"
        },
        "human_agreement_on_a_comparable_task": {
            "pair": agreement["pair"],
            "exact": agreement["exact_agreement_jaccard"],
            "loose": agreement["loose_agreement_share_one"],
            "meaning": (
                "Two experts publishing through NIST OLIR agreed on this share of their "
                "pairs. Any figure attached to an automated crosswalk should be read "
                "against it."
            ),
        },
        "attack_not_a_direct_target": (
            "DORA states duties on defenders and ATT&CK describes attacker behaviour. "
            "Matching them by text similarity produced confident nonsense, so the route "
            "to ATT&CK runs through 800-53: the first leg proposed, the second leg "
            "MITRE's published crosswalk."
        ),
        "candidates": rows,
        "composed_to_attack": composed,
    }

    REPORTS.mkdir(exist_ok=True)
    (REPORTS / "dora-candidates.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=1) + "\n", encoding="utf-8"
    )
    print(f"\n{len(rows)} candidates → {(REPORTS / 'dora-candidates.json').relative_to(ROOT)}")

    for name in TARGETS:
        subset = [r for r in rows if r["target_framework"] == name and r["rank"] == 1]
        best = sorted(subset, key=lambda r: -r["similarity"])[:3]
        worst = sorted(subset, key=lambda r: r["similarity"])[:2]
        print(f"\n── {name} ──")
        for r in best:
            print(
                f"   {r['similarity']:.2f}  {r['dora']:<16} → {r['target']}  {r['target_title'][:44]}"
            )
        print("   … weakest first-ranked:")
        for r in worst:
            print(
                f"   {r['similarity']:.2f}  {r['dora']:<16} → {r['target']}  {r['target_title'][:44]}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
